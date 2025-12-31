# -*- coding: utf-8 -*-
# train_export_model.py — Bulge η 제외, Tension η·εf은 입력, 5타깃 학습
import os, warnings
warnings.filterwarnings("ignore")

from pathlib import Path
import numpy as np
import pandas as pd
from openpyxl import load_workbook

from sklearn.model_selection import train_test_split
from sklearn.compose import ColumnTransformer
from sklearn.preprocessing import OneHotEncoder
from sklearn.pipeline import Pipeline
from sklearn.impute import SimpleImputer
from sklearn.metrics import r2_score, mean_absolute_error
from sklearn.ensemble import ExtraTreesRegressor
from sklearn.multioutput import MultiOutputRegressor

# ===== 설정 =====
DATA_XLSX = "250811_산학프로젝트_포스코의 워크시트.xlsx"
SHEET_FALLBACK = "학습DB"

# ★ 타깃(총 5개): Notch 2 + Shear 2 + Bulge εf (Bulge η 제외)
TARGETS_FIXED = [
    "Triaxiality(Notch R05)",     "Fracture strain(Notch R05)",
    "Triaxiality(Shear0)",        "Fracture strain(Shear0)",
    "Fracture strain(Punch Bulge)"
]

# ★ 입력 화이트리스트: Tension의 η, εf은 항상 입력(노란값)
FORCE_INPUT = {"Triaxiality(Tension)", "Fracture strain(Tension)"}

# ===== 유틸 =====
def read_excel_safe(path):
    ext = Path(path).suffix.lower()
    try:
        if ext in [".xlsx", ".xlsm", ".xltx", ".xltm"]:
            return pd.read_excel(path, engine="openpyxl")
        elif ext == ".xls":
            return pd.read_excel(path, engine="xlrd")
        elif ext == ".xlsb":
            return pd.read_excel(path, engine="pyxlsb")
    except Exception:
        pass
    return pd.read_excel(path)

def rgb_from_cell(cell):
    try:
        fill = cell.fill
        if fill is None or fill.fill_type is None: return None
        rgb = None
        if hasattr(fill, "fgColor") and getattr(fill.fgColor, "type", None) == "rgb":
            rgb = fill.fgColor.rgb
        if (rgb is None or rgb == "00000000") and hasattr(fill, "start_color"):
            rgb = getattr(fill.start_color, "rgb", None)
        if rgb is None: return None
        rgb = rgb.replace("0x","").replace("#","")
        if len(rgb)==8: rgb = rgb[2:]
        return rgb.upper() if len(rgb)==6 else None
    except:
        return None

def classify_rgb(rgb):
    if not rgb: return None
    r = int(rgb[0:2],16); g = int(rgb[2:4],16); b = int(rgb[4:6],16)
    if r>=200 and g>=200 and b<=120: return "yellow"   # 입력
    if r<=180 and g>=170 and b>=190: return "skyblue"  # 출력
    return None

def _as_series_single(X: pd.DataFrame, colname: str):
    """중복 컬럼이 있어도 단일 Series로 축약(왼쪽→오른쪽 첫 유효값)."""
    if colname not in X.columns:
        return None
    cols = [c for c in X.columns if c == colname]
    sub = X[cols]
    if isinstance(sub, pd.Series):
        return pd.to_numeric(sub, errors="coerce")
    sub_num = sub.apply(pd.to_numeric, errors="coerce")
    ser = sub_num.bfill(axis=1).iloc[:, 0]
    return ser

# =========================
# ✅ 핵심: pkl 저장 없이 번들 반환. seed=None이면 매번 랜덤 학습
# =========================
def train_bundle(data_xlsx: str = DATA_XLSX, seed: int | None = None):
    # 매 시행 랜덤 학습
    if seed is None:
        seed = int(np.random.randint(0, 2**31 - 1))

    # ===== 데이터 로드 =====
    wb = load_workbook(data_xlsx, data_only=True)
    sheet_name = wb.sheetnames[0] if wb.sheetnames else SHEET_FALLBACK
    ws = wb[sheet_name]
    df = read_excel_safe(data_xlsx)
    df = df.dropna(axis=1, how="all")

    # ===== 입력/출력 식별 =====
    headers, classes = [], []
    for col in range(1, ws.max_column+1):
        cell = ws.cell(row=1, column=col)
        headers.append(cell.value)
        classes.append(classify_rgb(rgb_from_cell(cell)))

    input_cols, output_cols = [], []
    for name, cls in zip(headers, classes):
        if name in df.columns:
            if cls == "yellow":   input_cols.append(name)
            elif cls == "skyblue": output_cols.append(name)

    # 패턴 보조
    def infer_by_name(columns):
        in_pats = ["yield","ultimate","elongation","r-value","tension","입력","input"]
        out_pats= ["notch","shear","bulge","punch","fracture","triaxiality","예측","output"]
        ins, outs = [], []
        for c in columns:
            cl = str(c).lower()
            if any(p in cl for p in in_pats): ins.append(c)
            if any(p in cl for p in out_pats): outs.append(c)
        return ins, outs

    if len(output_cols) == 0:
        inf_in, inf_out = infer_by_name(list(df.columns))
        output_cols = [c for c in inf_out if pd.api.types.is_numeric_dtype(df[c])]
        input_cols = [c for c in df.columns if c not in output_cols and pd.api.types.is_numeric_dtype(df[c])]

    cat_inputs = ['Material'] if 'Material' in df.columns else []

    # ★ tension 입력 화이트리스트 강제 반영
    for c in list(df.columns):
        if c in FORCE_INPUT:
            if c in output_cols:
                output_cols.remove(c)
            if pd.api.types.is_numeric_dtype(df[c]) and c not in input_cols:
                input_cols.append(c)

    # ★ 타깃은 TARGETS_FIXED(존재하는 것만)로 확정
    output_cols = [c for c in TARGETS_FIXED if c in df.columns]

    # 숫자 변환(학습 사용 컬럼만)
    for col in set(input_cols + output_cols):
        try:
            df[col] = pd.to_numeric(df[col], errors="coerce")
        except Exception:
            pass

    # ===== 특징공학 =====
    def build_enhanced_features(df_):
        X = df_[input_cols + cat_inputs].copy()

        ys  = _as_series_single(X, 'Yield Stress(MPa)')
        uts = _as_series_single(X, 'Ultimate Tensile Stress(MPa)')
        tel = _as_series_single(X, 'Total Elongation')
        rv  = _as_series_single(X, 'r-value')
        etaT= _as_series_single(X, 'Triaxiality(Tension)')
        efT = _as_series_single(X, 'Fracture strain(Tension)')

        if uts is not None and ys is not None:
            X['UTS_to_Yield']   = uts / (ys + 1e-8)
            X['Strength_Range'] = uts - ys
        if tel is not None:
            X['Elong_sq'] = tel**2
            X['Elong_sqrt'] = np.sqrt(tel + 1e-8)
            if rv is not None:
                X['Elong_x_r']   = tel * rv
                X['Elong_div_r'] = tel / (rv + 1e-8)
        if etaT is not None:
            X['etaT_sq']   = etaT**2
            X['etaT_cube'] = etaT**3
            if tel is not None:
                X['Total Elongation_x_Triaxiality(Tension)'] = tel * etaT
            if rv is not None:
                X['r-value_x_Triaxiality(Tension)'] = rv * etaT
        if rv is not None:
            X['r_sq']  = rv**2
            X['r_log'] = np.log(rv + 1e-8)
        if efT is not None:
            X['efT_sq'] = efT**2
            if etaT is not None:
                X['efT_x_etaT'] = efT * etaT

        # 상호작용
        key = [('Total Elongation', tel), ('r-value', rv), ('Triaxiality(Tension)', etaT)]
        for i in range(len(key)):
            for j in range(i+1, len(key)):
                n1,s1 = key[i]; n2,s2 = key[j]
                if s1 is not None and s2 is not None:
                    X[f'{n1}_x_{n2}'] = s1 * s2
        return X

    # ===== 모델 데이터 =====
    needed = list(dict.fromkeys(input_cols + cat_inputs + output_cols))
    df_model = df[needed].copy()
    df_model = df_model.dropna(subset=[c for c in output_cols if c in df_model.columns], how="any")

    X_all = build_enhanced_features(df_model)
    y_all = df_model[[c for c in output_cols if c in df_model.columns]].copy()

    # 분할
    X_train, X_test, y_train, y_test = train_test_split(
        X_all, y_all, test_size=0.25, random_state=seed, shuffle=True
    )

    # ===== 파이프라인 =====
    num_cols = [c for c in X_all.columns if c not in cat_inputs]
    try:
        ohe = OneHotEncoder(handle_unknown='ignore', sparse_output=False)
    except TypeError:
        ohe = OneHotEncoder(handle_unknown='ignore', sparse=False)

    pre = ColumnTransformer(
        transformers=[
            ('num', SimpleImputer(strategy='median'), num_cols),
            ('cat', ohe, cat_inputs)
        ],
        remainder='drop'
    )

    def create_et():
        return ExtraTreesRegressor(
            n_estimators=300, max_depth=None, max_features='sqrt',
            bootstrap=True, random_state=seed, n_jobs=-1
        )

    pipe = Pipeline([
        ('pre', pre),
        ('reg', MultiOutputRegressor(create_et()))
    ])

    pipe.fit(X_train, y_train)

    # 메트릭
    def compute_metrics(y_true, y_hat):
        d = {}
        for c in y_true.columns:
            d[c] = {
                "R2":  float(r2_score(y_true[c], y_hat[c])),
                "MAE": float(mean_absolute_error(y_true[c], y_hat[c])),
                "MAPE": float(np.mean(np.abs((y_true[c]-y_hat[c])/(y_true[c]+1e-8))) * 100),
            }
        return d

    y_pred = pd.DataFrame(pipe.predict(X_test), columns=y_test.columns, index=y_test.index)
    r2_flat = float(r2_score(y_test.values.flatten(), y_pred.values.flatten()))
    metrics = compute_metrics(y_test, y_pred)

    print("\n=== 학습 요약 ===")
    print(f"seed = {seed}")
    print(f"출력 타깃(5개 예상): {list(y_all.columns)}")
    print(f"전체 평탄화 R² = {r2_flat:.6f}")

    meta = {
        "input_cols": input_cols,
        "cat_inputs": cat_inputs,
        "output_cols": list(y_all.columns),
        "num_medians": df[input_cols].apply(pd.to_numeric, errors="coerce").median(numeric_only=True).to_dict(),
        "feature_columns": list(X_all.columns),
        "seed": seed
    }

    return {"model": pipe, "meta": meta, "metrics": metrics, "r2_flat": r2_flat}

if __name__ == "__main__":
    train_bundle()
