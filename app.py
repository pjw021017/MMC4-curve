# -*- coding: utf-8 -*-
"""
Streamlit app — 학습DB 엑셀에서 바로 학습해서 MMC4 curve를 그려주는 앱
- pkl 사용 안 함
- 실행 시 엑셀을 읽어 모델을 학습하고, 해당 모델을 이용해 Notch / Shear / Bulge를 예측
- 예측된 4개 파단점(Tension 포함)으로 MMC4 식의 C1~C6을 피팅
- 강종 분류(Mild / HSS / AHSS)에 따라 사전에 계산해 둔 밴드(10–90% 범위)를 함께 시각화
"""

import warnings
warnings.filterwarnings("ignore")

import os
from pathlib import Path

import numpy as np
import pandas as pd
import streamlit as st
import matplotlib.pyplot as plt

from openpyxl import load_workbook

from sklearn.compose import ColumnTransformer
from sklearn.preprocessing import OneHotEncoder
from sklearn.pipeline import Pipeline
from sklearn.impute import SimpleImputer
from sklearn.ensemble import ExtraTreesRegressor
from sklearn.multioutput import MultiOutputRegressor

from scipy.optimize import least_squares
from scipy.interpolate import CubicHermiteSpline

# =========================
# 기본 설정
# =========================
DATA_XLSX = "250811_산학프로젝트_포스코의 워크시트.xlsx"
SHEET_FALLBACK = "학습DB"

# 타깃(5개) : Notch 2 + Shear 2 + Bulge εf
TARGETS_FIXED = [
    "Triaxiality(Notch R05)",     "Fracture strain(Notch R05)",
    "Triaxiality(Shear0)",        "Fracture strain(Shear0)",
    "Fracture strain(Punch Bulge)"
]

# 항상 입력에 포함할 컬럼
FORCE_INPUT = {"Triaxiality(Tension)", "Fracture strain(Tension)"}

# Bulge η, 플로팅 범위
ETA_BULGE = 2.0 / 3.0
ETA_LO, ETA_HI = -0.1, 0.7

# 밴드 계산용 공통 η 축
ETA_GRID_BAND = np.linspace(ETA_LO, ETA_HI, 120)

st.set_page_config(page_title="MMC4 Predictor", layout="wide")
st.title("POSCO MMC4 — 모델 예측 기반 MMC4 곡선 시각화")

# =========================
# 세션 상태 초기화
# =========================
if "mmc4_ready" not in st.session_state:
    st.session_state["mmc4_ready"] = False

# =========================
# Sidebar: 재학습 트리거
# =========================
st.sidebar.header("학습")
st.sidebar.caption("모델/밴드는 앱 실행 시 엑셀에서 다시 학습됩니다.")
do_retrain = st.sidebar.button("모델 + 밴드 다시 학습")

# =========================
# 엑셀/색상 기반 유틸
# =========================
def read_excel_safe(path: str) -> pd.DataFrame:
    ext = Path(path).suffix.lower()
    try:
        if ext in [".xlsx", ".xlsm", ".xltx", ".xltm"]:
            return pd.read_excel(path, engine="openpyxl")
        elif ext == ".xls":
            return pd.read_excel(path, engine="xlrd")
        elif ext == ".xlsb":
            return pd.read_excel(path, engine="pyxlsb")
    except Exception:
        pass
    return pd.read_excel(path)


def rgb_from_cell(cell):
    try:
        fill = cell.fill
        if fill is None or fill.fill_type is None:
            return None
        rgb = None
        if hasattr(fill, "fgColor") and getattr(fill.fgColor, "type", None) == "rgb":
            rgb = fill.fgColor.rgb
        if (rgb is None or rgb == "00000000") and hasattr(fill, "start_color"):
            rgb = getattr(fill.start_color, "rgb", None)
        if rgb is None:
            return None
        rgb = rgb.replace("0x", "").replace("#", "")
        if len(rgb) == 8:
            rgb = rgb[2:]
        return rgb.upper() if len(rgb) == 6 else None
    except Exception:
        return None


def classify_rgb(rgb):
    if not rgb:
        return None
    r = int(rgb[0:2], 16)
    g = int(rgb[2:4], 16)
    b = int(rgb[4:6], 16)
    if r >= 200 and g >= 200 and b <= 120:
        return "yellow"
    if r <= 180 and g >= 170 and b >= 190:
        return "skyblue"
    return None


def infer_by_name(columns):
    in_pats = ["yield", "ultimate", "elongation", "r-value", "tension", "입력", "input"]
    out_pats = ["notch", "shear", "bulge", "punch", "fracture", "triaxiality", "예측", "output"]
    ins, outs = [], []
    for c in columns:
        cl = str(c).lower()
        if any(p in cl for p in in_pats):
            ins.append(c)
        if any(p in cl for p in out_pats):
            outs.append(c)
    return ins, outs


def _as_series_single(X: pd.DataFrame, colname: str):
    if colname not in X.columns:
        return None
    cols = [c for c in X.columns if c == colname]
    sub = X[cols]
    if isinstance(sub, pd.Series):
        return pd.to_numeric(sub, errors="coerce")
    sub_num = sub.apply(pd.to_numeric, errors="coerce")
    ser = sub_num.bfill(axis=1).iloc[:, 0]
    return ser


# =========================
# 강종 분류 유틸
# =========================
def classify_steel_by_yield(ys: float) -> str:
    """Yield stress 기준으로 Mild / HSS / AHSS 분류.
    - Mild : YS < 300 MPa
    - HSS  : 300 ≤ YS < 600 MPa
    - AHSS : YS ≥ 600 MPa
    필요하면 실제 DB 기준에 맞게 값만 수정해서 쓰면 된다.
    """
    if ys is None or np.isnan(ys):
        return "Unknown"
    if ys < 300.0:
        return "Mild"
    elif ys < 600.0:
        return "HSS"
    else:
        return "AHSS"


# =========================
# 특징공학
# =========================
def build_enhanced_features(df_, input_cols, cat_inputs):
    X = df_[input_cols + cat_inputs].copy()

    ys = _as_series_single(X, 'Yield Stress(MPa)')
    uts = _as_series_single(X, 'Ultimate Tensile Stress(MPa)')
    tel = _as_series_single(X, 'Total Elongation')
    rv = _as_series_single(X, 'r-value')
    etaT = _as_series_single(X, 'Triaxiality(Tension)')
    efT = _as_series_single(X, 'Fracture strain(Tension)')

    if uts is not None and ys is not None:
        X['UTS_to_Yield'] = uts / (ys + 1e-8)
        X['Strength_Range'] = uts - ys
    if tel is not None:
        X['Elong_sq'] = tel ** 2
        X['Elong_sqrt'] = np.sqrt(tel + 1e-8)
        if rv is not None:
            X['Elong_x_r'] = tel * rv
            X['Elong_div_r'] = tel / (rv + 1e-8)
    if etaT is not None:
        X['etaT_sq'] = etaT ** 2
        X['etaT_cube'] = etaT ** 3
        if tel is not None:
            X['Total Elongation_x_Triaxiality(Tension)'] = tel * etaT
        if rv is not None:
            X['r-value_x_Triaxiality(Tension)'] = rv * etaT
    if rv is not None:
        X['r_sq'] = rv ** 2
        X['r_log'] = np.log(rv + 1e-8)
    if efT is not None:
        X['efT_sq'] = efT ** 2
        if etaT is not None:
            X['efT_x_etaT'] = efT * etaT

    key = [('Total Elongation', tel),
           ('r-value', rv),
           ('Triaxiality(Tension)', etaT)]
    for i in range(len(key)):
        for j in range(i + 1, len(key)):
            n1, s1 = key[i]
            n2, s2 = key[j]
            if s1 is not None and s2 is not None:
                X[f'{n1}_x_{n2}'] = s1 * s2

    return X


# =========================
# MMC4 수식/피팅
# =========================
def theta_bar(eta):
    eta = np.asarray(eta, dtype=float)
    arg = -(27.0 / 2.0) * eta * (eta ** 2 - 1.0 / 3.0)
    arg = np.clip(arg, -1.0, 1.0)
    return 1.0 - (2.0 / np.pi) * np.arccos(arg)


def c6_effective(eta, C6):
    t = 1.0 / np.sqrt(3.0)
    eta = np.asarray(eta, dtype=float)
    return np.where((eta <= -t) | ((eta >= 0) & (eta <= t)), 1.0, float(C6))


def mmc4_eps(eta, C):
    C1, C2, C3, C4, C5, C6 = C
    eta = np.asarray(eta, dtype=float)

    tb = theta_bar(eta)
    c6e = c6_effective(eta, C6)
    k = np.sqrt(3.0) / (2.0 - np.sqrt(3.0))

    term1 = (C1 / C4) * (C5 + k * (c6e - C5) * (1.0 / np.cos(tb * np.pi / 6.0) - 1.0))

    base = np.sqrt(1.0 + (C3 ** 2) / 3.0) * np.cos(tb * np.pi / 6.0) \
           + C3 * (eta + (1.0 / 3.0) * np.sin(tb * np.pi / 6.0))
    base = np.maximum(base, 1e-6)

    return term1 * (base ** (-1.0 / C2))


def fit_mmc4(etas, epss):
    etas = np.asarray(etas, dtype=float)
    epss = np.asarray(epss, dtype=float)

    def resid(p):
        return mmc4_eps(etas, p) - epss

    x0 = np.array([1.0, 1.0, 0.2, 1.0, 0.6, 0.8])
    lb = np.array([0.001, 0.10, -2.0, 0.10, 0.0, 0.0])
    ub = np.array([10.0, 5.0,  2.0, 5.0,  2.0, 2.0])

    res = least_squares(resid, x0, bounds=(lb, ub), max_nfev=5000, verbose=0)
    return res.x


def extend_curve_with_hermite(eta_grid, C):
    """Shear~Bulge 구간은 MMC4 그대로, Bulge 이후는 Hermite 외삽."""
    eps_curve = mmc4_eps(eta_grid, C)

    # Bulge 접점에서의 값/기울기
    h = 1e-4
    e_max = float(mmc4_eps(ETA_BULGE, C))
    d_max = float((mmc4_eps(ETA_BULGE + h, C) - mmc4_eps(ETA_BULGE - h, C)) / (2 * h))

    y_hi = e_max + d_max * (ETA_HI - ETA_BULGE)
    right_conn = CubicHermiteSpline([ETA_BULGE, ETA_HI], [e_max, y_hi], [d_max, d_max])

    mask_right = eta_grid > ETA_BULGE
    if np.any(mask_right):
        eps_curve[mask_right] = right_conn(eta_grid[mask_right])

    return eps_curve, e_max, d_max, right_conn


# =========================
# 모델 학습 (엑셀 → 파이프라인 + 강종별 밴드)
# =========================
@st.cache_resource
def load_and_train():
    if not os.path.exists(DATA_XLSX):
        raise FileNotFoundError(f"엑셀 파일을 찾을 수 없음: {DATA_XLSX}")

    # 엑셀 로드 (색상 정보용)
    wb = load_workbook(DATA_XLSX, data_only=True)
    sheet_name = wb.sheetnames[0] if wb.sheetnames else SHEET_FALLBACK
    ws = wb[sheet_name]

    df = read_excel_safe(DATA_XLSX)
    df = df.dropna(axis=1, how="all")

    # 헤더 색상 파싱
    headers, classes = [], []
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col)
        headers.append(cell.value)
        classes.append(classify_rgb(rgb_from_cell(cell)))

    input_cols, output_cols = [], []
    for name, cls in zip(headers, classes):
        if name in df.columns:
            if cls == "yellow":
                input_cols.append(name)
            elif cls == "skyblue":
                output_cols.append(name)

    # 색상 정보가 없으면 컬럼명으로 추론
    if len(output_cols) == 0:
        inf_in, inf_out = infer_by_name(list(df.columns))
        output_cols = [c for c in inf_out if pd.api.types.is_numeric_dtype(df[c])]
        input_cols = [c for c in df.columns if c not in output_cols and pd.api.types.is_numeric_dtype(df[c])]

    cat_inputs = ["Material"] if "Material" in df.columns else []

    # Tension η, εf 는 항상 입력으로 강제
    for c in list(df.columns):
        if c in FORCE_INPUT:
            if c in output_cols:
                output_cols.remove(c)
            if pd.api.types.is_numeric_dtype(df[c]) and c not in input_cols:
                input_cols.append(c)

    # 타깃은 고정 리스트 중 실제 존재하는 것만 사용
    output_cols = [c for c in TARGETS_FIXED if c in df.columns]

    # 숫자 변환
    for col in set(input_cols + output_cols):
        try:
            df[col] = pd.to_numeric(df[col], errors="coerce")
        except Exception:
            pass

    needed = list(dict.fromkeys(input_cols + cat_inputs + output_cols))
    df_model = df[needed].copy()
    df_model = df_model.dropna(subset=[c for c in output_cols if c in df_model.columns], how="any")

    # ---------------- MMC4 곡선으로부터 강종별 밴드 계산 ----------------
    curves_by_class = {"Mild": [], "HSS": [], "AHSS": []}

    # 밴드 계산에 필요한 컬럼
    cols_req = [
        "Yield Stress(MPa)",
        "Triaxiality(Shear0)", "Fracture strain(Shear0)",
        "Triaxiality(Tension)", "Fracture strain(Tension)",
        "Triaxiality(Notch R05)", "Fracture strain(Notch R05)",
        "Fracture strain(Punch Bulge)"
    ]
    missing_cols = [c for c in cols_req if c not in df_model.columns]
    if len(missing_cols) == 0:
        for _, row in df_model.iterrows():
            ys = row["Yield Stress(MPa)"]
            steel_class = classify_steel_by_yield(ys)
            if steel_class not in curves_by_class:
                continue

            vals = [row["Triaxiality(Shear0)"], row["Fracture strain(Shear0)"],
                    row["Triaxiality(Tension)"], row["Fracture strain(Tension)"],
                    row["Triaxiality(Notch R05)"], row["Fracture strain(Notch R05)"],
                    row["Fracture strain(Punch Bulge)"]]
            if any(pd.isna(v) for v in vals):
                continue

            pts = [
                ("Shear",   float(row["Triaxiality(Shear0)"]),     float(row["Fracture strain(Shear0)"])),
                ("Tension", float(row["Triaxiality(Tension)"]),   float(row["Fracture strain(Tension)"])),
                ("Notch",   float(row["Triaxiality(Notch R05)"]), float(row["Fracture strain(Notch R05)"])),
                ("Bulge",   float(ETA_BULGE),                     float(row["Fracture strain(Punch Bulge)"])),
            ]
            pts = sorted(pts, key=lambda t: t[1])
            etas = np.array([p[1] for p in pts], dtype=float)
            epss = np.array([p[2] for p in pts], dtype=float)

            try:
                C_row = fit_mmc4(etas, epss)
                eps_curve_row, _, _, _ = extend_curve_with_hermite(ETA_GRID_BAND, C_row)
                curves_by_class[steel_class].append(eps_curve_row)
            except Exception:
                continue

    bands = {}
    for cls_name, curves in curves_by_class.items():
        if len(curves) == 0:
            continue
        arr = np.vstack(curves)  # [n_samples, n_eta]
        lo = np.nanpercentile(arr, 10, axis=0)
        hi = np.nanpercentile(arr, 90, axis=0)
        bands[cls_name] = {"lo": lo, "hi": hi}

    # ---------------- 예측 모델 학습 ----------------
    X_all = build_enhanced_features(df_model, input_cols, cat_inputs)
    y_all = df_model[[c for c in output_cols if c in df_model.columns]].copy()

    num_cols = [c for c in X_all.columns if c not in cat_inputs]
    try:
        ohe = OneHotEncoder(handle_unknown="ignore", sparse_output=False)
    except TypeError:
        ohe = OneHotEncoder(handle_unknown="ignore", sparse=False)

    pre = ColumnTransformer(
        transformers=[
            ("num", SimpleImputer(strategy="median"), num_cols),
            ("cat", ohe, cat_inputs),
        ],
        remainder="drop",
    )

    def create_et():
        return ExtraTreesRegressor(
            n_estimators=300,
            max_depth=None,
            max_features="sqrt",
            bootstrap=True,
            random_state=None,
            n_jobs=-1,
        )

    pipe = Pipeline([
        ("pre", pre),
        ("reg", MultiOutputRegressor(create_et())),
    ])
    pipe.fit(X_all, y_all)

    meta = {
        "input_cols": input_cols,
        "cat_inputs": cat_inputs,
        "output_cols": list(y_all.columns),
        "num_medians": df[input_cols].apply(pd.to_numeric, errors="coerce").median(numeric_only=True).to_dict(),
        "feature_columns": list(X_all.columns),
        "bands": bands,
        "eta_grid_band": ETA_GRID_BAND,
    }
    return pipe, meta


# 재학습 트리거 처리
if do_retrain:
    load_and_train.clear()
    st.session_state["mmc4_ready"] = False
    for k in ["pts", "C_hat", "ef_bulge", "e_max", "d_max", "steel_class"]:
        st.session_state.pop(k, None)
    st.rerun()


# =========================
# 모델 로드
# =========================
try:
    model, meta = load_and_train()
except Exception as e:
    st.error(f"모델/밴드 학습 실패: {e}")
    st.stop()

INPUTS = meta["input_cols"]
CAT_INPUTS = meta["cat_inputs"]
OUTPUTS = meta["output_cols"]
MEDIANS = meta.get("num_medians", {})
BANDS = meta.get("bands", {})
ETA_GRID_BAND = meta.get("eta_grid_band", ETA_GRID_BAND)


# =========================
# 입력 폼
# =========================
st.header("입력값(노란색)")
c1, c2 = st.columns(2)

with c1:
    ys = st.number_input(
        "Yield Stress(MPa)",
        value=float(MEDIANS.get("Yield Stress(MPa)", 200.0)),
    )
    uts = st.number_input(
        "Ultimate Tensile Stress(MPa)",
        value=float(MEDIANS.get("Ultimate Tensile Stress(MPa)", 350.0)),
    )
    tel = st.number_input(
        "Total Elongation",
        value=float(MEDIANS.get("Total Elongation", 0.20)),
        format="%.5f",
    )

with c2:
    rv = st.number_input(
        "r-value",
        value=float(MEDIANS.get("r-value", 1.00)),
        format="%.5f",
    )
    etaT = st.number_input(
        "Triaxiality(Tension)",
        value=float(MEDIANS.get("Triaxiality(Tension)", 0.33)),
        min_value=-0.5,
        max_value=0.7,
        step=0.01,
        format="%.2f",
    )
    efT = st.number_input(
        "Fracture strain(Tension)",
        value=0.20,
        min_value=0.0,
        step=0.001,
        format="%.5f",
    )

extra_inputs = [
    c for c in INPUTS
    if c not in [
        "Yield Stress(MPa)", "Ultimate Tensile Stress(MPa)",
        "Total Elongation", "r-value",
        "Triaxiality(Tension)", "Fracture strain(Tension)",
    ]
]
adv_vals = {}
if len(extra_inputs) or len(CAT_INPUTS):
    with st.expander("고급 옵션", expanded=False):
        for c in CAT_INPUTS:
            adv_vals[c] = st.text_input(c, value="")
        for c in extra_inputs:
            adv_vals[c] = st.number_input(c, value=float(MEDIANS.get(c, 0.0)))

# =========================
# 예측 및 MMC4 플롯
# =========================
if st.button("예측 및 MMC4 플롯"):
    # 입력 한 줄 구성
    x_dict = {
        "Yield Stress(MPa)": ys,
        "Ultimate Tensile Stress(MPa)": uts,
        "Total Elongation": tel,
        "r-value": rv,
        "Triaxiality(Tension)": etaT,
        "Fracture strain(Tension)": efT,
    }
    for c in extra_inputs:
        x_dict[c] = adv_vals.get(c, np.nan)
    for c in CAT_INPUTS:
        x_dict[c] = adv_vals.get(c, "")

    x_one = pd.DataFrame([x_dict])

    for col in INPUTS:
        if col not in x_one.columns:
            x_one[col] = np.nan
    for col in CAT_INPUTS:
        if col not in x_one.columns:
            x_one[col] = ""

    X_tmp = build_enhanced_features(x_one, INPUTS, CAT_INPUTS)

    pre = model.named_steps["pre"]
    num_cols_model = list(next(t[2] for t in pre.transformers if t[0] == "num"))
    cat_cols_model = []
    for name, trans, cols in pre.transformers:
        if name == "cat" and cols is not None and cols != "drop":
            cat_cols_model = list(cols)

    for c in num_cols_model:
        if c not in X_tmp.columns:
            X_tmp[c] = np.nan
    for c in cat_cols_model:
        if c not in x_one.columns:
            x_one[c] = ""

    X_feed = pd.DataFrame()
    X_feed[num_cols_model] = X_tmp[num_cols_model]
    if cat_cols_model:
        X_feed[cat_cols_model] = x_one[cat_cols_model].astype(str)

    y_hat = pd.Series(model.predict(X_feed)[0], index=OUTPUTS)

    # 4점 구성
    eta_shear = y_hat.get("Triaxiality(Shear0)")
    ef_shear = y_hat.get("Fracture strain(Shear0)")
    eta_notch = y_hat.get("Triaxiality(Notch R05)")
    ef_notch = y_hat.get("Fracture strain(Notch R05)")
    ef_bulge = y_hat.get("Fracture strain(Punch Bulge)")

    missing = []
    if eta_shear is None: missing.append("η(Shear)")
    if ef_shear is None: missing.append("εf(Shear)")
    if eta_notch is None: missing.append("η(Notch)")
    if ef_notch is None: missing.append("εf(Notch)")
    if ef_bulge is None: missing.append("εf(Bulge)")
    if missing:
        st.error("부족한 값: " + ", ".join(missing) + " — 엑셀/학습 스키마를 확인하세요.")
        st.stop()

    pts = [
        ("Shear",   float(eta_shear), float(ef_shear)),
        ("Tension", float(etaT),      float(efT)),
        ("Notch",   float(eta_notch), float(ef_notch)),
        ("Bulge",   float(ETA_BULGE), float(ef_bulge)),
    ]
    pts = sorted(pts, key=lambda t: t[1])

    etas = np.array([p[1] for p in pts], dtype=float)
    epss = np.array([p[2] for p in pts], dtype=float)
    C_hat = fit_mmc4(etas, epss)

    eps_curve, e_max, d_max, right_conn = extend_curve_with_hermite(ETA_GRID_BAND, C_hat)

    steel_class = classify_steel_by_yield(ys)

    st.session_state["mmc4_ready"] = True
    st.session_state["pts"] = pts
    st.session_state["C_hat"] = C_hat
    st.session_state["ef_bulge"] = float(ef_bulge)
    st.session_state["e_max"] = e_max
    st.session_state["d_max"] = d_max
    st.session_state["steel_class"] = steel_class
    st.session_state["eps_curve"] = eps_curve
    st.session_state["right_conn"] = right_conn


# =========================
# 결과 표시
# =========================
if st.session_state.get("mmc4_ready", False):
    pts = st.session_state["pts"]
    C_hat = st.session_state["C_hat"]
    ef_bulge = st.session_state["ef_bulge"]
    steel_class = st.session_state.get("steel_class", "Unknown")
    eps_curve = st.session_state["eps_curve"]

    # 플롯
    fig, ax = plt.subplots(figsize=(7, 4.5))

    # 밴드 색상 정의
    band_colors = {
        "Mild": "#FFEFC2",   # 연한 노랑
        "HSS": "#D5F0C9",    # 연한 초록
        "AHSS": "#CBE4FF",   # 연한 파랑
    }

    # 밴드: Mild, HSS, AHSS 순서
    for cls_name in ["Mild", "HSS", "AHSS"]:
        band = BANDS.get(cls_name)
        if band is None:
            continue
        lo = band["lo"]
        hi = band["hi"]
        color = band_colors.get(cls_name, "0.9")
        label = f"{cls_name} band (10–90%)"
        ax.fill_between(ETA_GRID_BAND, lo, hi, color=color, alpha=0.35, label=label)

    # MMC4 커브 (현재 입력 강종, 진한 선)
    ax.plot(ETA_GRID_BAND, eps_curve, lw=2.5, color="tab:green", label="MMC4 curve (fit)")

    # 4점
    marker_colors = {
        "Shear": "tab:blue",
        "Tension": "tab:orange",
        "Notch": "tab:green",
        "Bulge": "tab:red",
    }
    for name, x, y in pts:
        ax.scatter([x], [y], s=60, edgecolor="k", zorder=5,
                   color=marker_colors.get(name, "k"), label=name)
        ax.annotate(name, (x, y), xytext=(5, 5), textcoords="offset points", fontsize=9)

    ax.set_xlabel("Triaxiality (η)")
    ax.set_ylabel("Fracture strain (εf)")
    ax.set_title(f"MMC4 Curve  |  Steel class: {steel_class}")
    ax.set_xlim(ETA_LO, ETA_HI)
    ax.set_ylim(0.0, float(ef_bulge) + 0.3)
    ax.grid(True, alpha=0.3)
    ax.legend(loc="upper center", bbox_to_anchor=(0.5, -0.18), ncol=3, frameon=False)

    plt.tight_layout()
    st.pyplot(fig)

    # 4점/파라미터 출력
    st.markdown("**4점 요약**")
    for name, x, y in pts:
        st.text(f"{name:8s} η={x:.4f}  εf={y:.5f}")

    st.markdown("**적합 파라미터 C1~C6**")
    st.text("C = [" + ", ".join(f"{v:.6f}" for v in C_hat) + "]")

    # 특정 η → εf 조회
    st.markdown("**특정 η 입력 → εf 출력**")
    eta_query = st.number_input(
        "조회할 Triaxiality (η_query)",
        value=float(st.session_state.get("eta_query", 0.30000)),
        min_value=float(ETA_LO),
        max_value=float(ETA_HI),
        step=0.01,
        format="%.5f",
        key="eta_query",
    )

    if eta_query <= ETA_BULGE:
        eps_query = float(mmc4_eps(eta_query, C_hat))
    else:
        # Hermite 외삽 곡선은 eps_curve에 이미 반영되어 있으므로 재계산 없이 보간
        idx = np.searchsorted(ETA_GRID_BAND, eta_query)
        if idx <= 0:
            eps_query = float(eps_curve[0])
        elif idx >= len(ETA_GRID_BAND):
            eps_query = float(eps_curve[-1])
        else:
            x0, x1 = ETA_GRID_BAND[idx-1], ETA_GRID_BAND[idx]
            y0, y1 = eps_curve[idx-1], eps_curve[idx]
            t = (eta_query - x0) / (x1 - x0)
            eps_query = float(y0 + t * (y1 - y0))

    st.write(f"결과: η = {eta_query:.5f} 에서 εf = {eps_query:.6f} 입니다.")
else:
    st.info("입력값을 설정한 뒤, '예측 및 MMC4 플롯' 버튼을 눌러 결과를 생성하세요.")
